from openpyxl import Workbook,load_workbook
from openpyxl.utils import rows_from_range
from openpyxl.styles import Protection, PatternFill
#创建数据
def create_excel():
    wb=Workbook()
    sh=wb.active
    rows=[
        ['工号', '姓名', '部门', '基本工资', '提成', '加班工资', '社保扣除', '考勤扣除', '应发工资', '邮箱'],
          [10000, 'A', '甲', 8000, 7000, 1000, 680, 0, 15320, '1@163.com'],
          [10001, 'B', '乙', 4000, 3000, 1000, 680, 0, 7320, '1@164.com'],
          [10002, 'C', '丙', 6000, 5000, 1000, 680, 300, 11020, '1@165.com'],
          [10003, 'D', '丁', 7000, 6000, 1000, 680, 600, 12720, '1@166.com'],
          [10004, 'E', '戊', 5000, 4000, 1000, 680, 0, 9320, '1@167.com']
    ]
    for row in rows:
        sh.append(row)
    #修改样式
    bak_color=PatternFill('solid',fgColor='AEEEEE')
    for i in range(1,sh.max_row+1):
        if i%2==0:#如果被2整除为真，就是偶数
            for cell in range(1,sh.max_column+1):
                sh.cell(i,cell).fill=bak_color
                # print(sh.cell(i,cell))
    wb.save('./excel/工资表分成个人表格数据/隔行变色.xlsx')

if __name__ == '__main__':
    create_excel()
