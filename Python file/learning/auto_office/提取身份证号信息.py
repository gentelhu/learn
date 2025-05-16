from openpyxl import load_workbook
from datetime import datetime

def create_time():
    wb=load_workbook('./excel/person.xlsx')
    sh=wb.active
    now_year=datetime.now().year #获取当前年
    max_column=sh.max_column  #获取最后一列的数字
    for i,cell in enumerate(sh['c'],1):
        per=cell.value
        year=per[6:10]
        mouth=per[10:12]
        day=per[12:14]
        # print(year,mouth,day)
        age=now_year-int(year) #计算年龄
        sh.cell(i,max_column+1).value=year
        sh.cell(i,max_column+2).value=mouth
        sh.cell(i,max_column+3).value=day
        sh.cell(i,max_column+4).value=age
    wb.save('./excel/提取身份证信息.xlsx')

if __name__ == '__main__':
    create_time()
