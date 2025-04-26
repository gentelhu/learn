from openpyxl import load_workbook
from pandas.core.config_init import pc_max_info_rows_doc


#打开文件函数
def open():
    wb=load_workbook('./excel/bead_plan.xlsx')
    pcm=wb['PCM']
    print(pcm)
#显示表格里有哪些子文档
def show_sheels():
    wb = load_workbook('./excel/bead_plan.xlsx')
    print(wb.sheetnames)
    for sh in wb:
        print(sh.title)
#获取一个单元格
def get_one_value():
    wb = load_workbook('./excel/bead_plan.xlsx')
    pcm = wb['PCM']
    value1 = pcm['b2']
    print(value1)

#获取多个单元格
def get_many_value():
    wb = load_workbook('./excel/bead_plan.xlsx')
    pcm = wb['PCM']
    cells1 = pcm['b2':'c6']#获取多个单元格
    # print(cells1)
    #整行整列
    cells_row2= pcm[2]#整行
    cells_col2= pcm['b']#整列
    # print(cells_row2)
    # print(cells_col2)
    #多行多列
    cells_rwo2_5=pcm[2:5]#2到5行
    # print(cells_rwo2_5)

    #通过迭代获取几行到几行之间，前几列数据
    # for row in pcm.iter_rows(min_row=6,max_row=9,max_col=3):
    #     for cell in row:
            # print(cell.value,'\t',end='')
        # print()

    # 通过迭代获取几行到几行之间，几列到几列之间的数据
    for row in pcm.iter_rows(min_row=6, max_row=9,min_col=2,max_col=2):
        for cell in row:
            print(cell.value,'\t',end='')
        print()
#获取所有数据
def get_all_data():
    wb = load_workbook('./excel/bead_plan.xlsx')
    pcm = wb['PCM']
    #获取表格中所有行
    # for row in pcm.rows:
    #     for cell in row:
    #         print(cell.value,'\t',end='')
    #     print()

    # 获取表格中所有列
    for cloumn in pcm.columns:
        for cell in cloumn:
            print(cell.value, '\t', end='')
        print()
#获取表格中有多少行多少列
def get_num():
    wb = load_workbook('./excel/bead_plan.xlsx')
    pcm = wb['PCM']
    print(pcm.max_row)
    print(pcm.max_column)

if __name__ == '__main__':
    # open()
    # show_sheels()
    # get_one_value()
    # get_many_value()
    # get_all_data()
    get_num()

