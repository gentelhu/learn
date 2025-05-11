from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,colors

def new():
    wb=Workbook()
    sh1=wb.active#默认激活创建好的sheet1
    ss2=wb.create_sheet('数据')#创建一个数据表
    ss2=wb.create_sheet('人员',0)#将人员数据放在最前面
    wb.save('./excel/create.xlsx')

def set_value():
    hold_italic_30_font=Font(name='宋体',size=30,italic=True,bold=True,color=colors.BLUE)
    #italic斜体 bold加粗
    wb=Workbook()
    sh1=wb.active
    sh1['a1']='hello'
    sh1['b2']='python'
    sh1['b2'].font=hold_italic_30_font
    wb.save('./excel/create.xlsx')

def set2_value():
    wb=Workbook()
    sh1=wb.active
    hold_italic_30_font = Font(name='宋体', size=30, italic=True, bold=True, color=colors.BLUE)
    data=['python','hello','world']
    for i,d in enumerate(data):
        sh1.cell(i+1,1).value=d
        sh1.cell(i+1,1).font= hold_italic_30_font
    wb.save('./excel/create.xlsx')

def set_merge():
    wb=Workbook()
    sh1=wb.active
    sh1.merge_cells('a1:c1')#合并单元格
    sh1.merge_cells('d2:e5')
    sh1['a1']='横向合并'
    sh1['d2']='纵向合并'
    wb.save('./excel/create.xlsx')

if __name__ == '__main__':
    # new()
    # set_value()
    # set2_value()
    set_merge()