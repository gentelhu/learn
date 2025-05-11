from openpyxl import Workbook,load_workbook
from datetime import date

def create_data():
    wb=Workbook()
    sh=wb.active
    rows=[
        ['data','name','time'],
        [date(2020,12,1), 'tim', '18:58'],
        [date(2020, 12, 2), 'tom', '18:50'],
        [date(2020, 12, 3), 'jim', '18:20'],
        [date(2020, 12, 4), 'gentel', '19:02'],
    ]
    for row in rows:
        sh.append(row)
    wb.save('./excel/打卡时间.xlsx')

def statistics():
    #读取数据
    wb=load_workbook('./excel/打卡时间.xlsx')
    sh=wb.active
    data=[]
    for i in range(2,sh.max_row+1):
        t_data=[]
        for j in range(1,sh.max_column+1):
            t_data.append(sh.cell(i,j).value)
        #统计
        h,m=t_data[2].split(":")
        full = int(h)*60+int(m)
        tmp= full-18*60
        t_data.append(tmp)
        #处理时间
        t_data[0]=t_data[0].date()
        data.append(t_data)

    #保存
    n_wb=Workbook()
    n_sh=n_wb.create_sheet('统计加班')
    for d in data:
        n_sh.append(d)
    n_wb.save('./excel/统计加班时间.xlsx')


if __name__ == '__main__':
    # create_data()
    statistics()
