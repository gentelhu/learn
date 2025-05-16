from openpyxl import load_workbook

def sums():
    # 加载工作簿和工作表
    wb = load_workbook('./excel/bead_plan.xlsx')
    ws = wb['PCM']

    # 获取C列数据范围（假设数据从第2行开始，跳过表头）
    c_column = ws['C'][1:]  # 索引1表示从第2行开始

    # 计算C列数据总和
    total = sum(cell.value for cell in c_column if cell.value is not None)

    # 获取当前最后一行的行号
    last_row = ws.max_row

    # 将总和写入C列最后一行下方的单元格
    ws[f'C{last_row + 1}'] = total
    ws[f'B{last_row + 1}'] = '合计'

    # 保存工作簿
    wb.save('./excel/bead_plan.xlsx')

    print(f"C列数据总和: {total}")
    print(f"已将总和写入C列第{last_row + 1}行")

if __name__ == '__main__':
    sums()