from openpyxl import load_workbook,Workbook
import os
# print(os.listdir('./excel'))#显示目录下面有哪些文件
def copy_data():
    wb =Workbook()
    for name in os.listdir('./excel/合并数据'):
        path=f'./excel/合并数据/{name}'
        tmp_wb=load_workbook(path)
        tmp_sh=tmp_wb.active
        sh=wb.create_sheet(name[:-5])#文件名会带后缀xlsx,通过切片把后面的后.xlsx去掉
        for r in range(1,tmp_sh.max_row+1):
            #获取整行数据
            row_value=[]
            for c in range(1,tmp_sh.max_column+1):
                value=tmp_sh.cell(r,c).value
                row_value.append(value)
            sh.append(row_value)
    del wb['Sheet']#删除sheet这个表格
    wb.save('./excel/合并数据/合并excel成多个工作表.xlsx')
if __name__ == '__main__':
    copy_data()
