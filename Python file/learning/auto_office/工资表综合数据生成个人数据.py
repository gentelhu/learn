from numpy.ma.core import append
from openpyxl import load_workbook,Workbook

def create_excel():
    wb=load_workbook('./excel/工资表分成个人表格数据/工资条.xlsx',data_only=True)
    sh=wb.active
    title=['工号','姓名','部门','基本工资','提成','加班工资','社保扣除','考勤扣除','应发工资','邮箱']
    for i,row in enumerate(sh.rows):
        if i==0:
            continue
        else:
            temp_wb=Workbook()
            temp_sh=temp_wb.active
            temp_sh.append((title))
            row_data=[cell.value for cell in row]
            temp_sh.append(row_data)
            temp_wb.save(f'./excel/工资表分成个人表格数据/{row_data[1]}.xlsx')
            # print(row_data)
if __name__ == '__main__':
    create_excel()
