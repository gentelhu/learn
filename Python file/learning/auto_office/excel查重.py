from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill


def dum():
    wb=load_workbook('./excel/打卡时间.xlsx')
    sh=wb.active
    index=[] #存储哪一行是重复数据
    tmp=[]  #没有重复的数据
    for i,c in enumerate(sh['b'],1):
        flag=c.value not in tmp
        # print(flag,f'==={c}==={tmp}')
        # print(i)
        if flag:
            tmp.append(c.value)
        else:
            index.append(i)
    # print(tmp,index)
    fill=PatternFill('solid',fgColor='AEEEEE')
    for i,r in enumerate(sh.rows,1):
        if i in index:
            for c in r:
                c.fill=fill
            # print(f'第{i}条数据是重数据')
    wb.save('./excel/去重.xlsx')


if __name__ == '__main__':
    dum()


